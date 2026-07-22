"""분석 결과 내보내기: CSV / JSONL / Excel.

리뷰 본문과 AI 분석 결과를 한 행에 담아, 스프레드시트에서 바로 볼 수 있게 한다.
필터(감정·별점·기간·제품)는 list/dashboard 와 같은 조건 체계를 쓴다.
"""

from __future__ import annotations

import csv
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from src import config as config_module
from src.db import Database

logger = logging.getLogger(__name__)

# 내보내기 컬럼 순서(헤더는 한국어로 두어 스프레드시트에서 바로 읽히게 한다)
COLUMNS = [
    ("id", "ID"),
    ("review_date", "작성일"),
    ("product", "제품"),
    ("category", "카테고리"),
    ("rating", "별점"),
    ("sentiment", "감정"),
    ("score", "신뢰도"),
    ("keywords", "키워드"),
    ("lang", "언어"),
    ("text_len", "글자수"),
    ("review_text", "리뷰내용"),
    ("sentiment_model", "분석모델"),
    ("analyzed_at", "분석일시"),
]


def rows_to_dicts(rows) -> list[dict[str, Any]]:
    """DB 행을 내보내기용 딕셔너리로 바꾼다."""
    result = []
    for row in rows:
        result.append({header: row[key] for key, header in COLUMNS})
    return result


def export_csv(records: list[dict[str, Any]], path: Path) -> Path:
    """CSV 저장. Excel 한글 깨짐을 막기 위해 utf-8-sig(BOM) 로 쓴다."""
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=[header for _, header in COLUMNS])
        writer.writeheader()
        writer.writerows(records)
    return path


def export_jsonl(records: list[dict[str, Any]], path: Path) -> Path:
    """JSONL 저장. 한 줄에 리뷰 하나 — 스트리밍 처리에 유리하다."""
    with path.open("w", encoding="utf-8") as f:
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
    return path


def export_xlsx(records: list[dict[str, Any]], path: Path) -> Path:
    """Excel 저장. 헤더 고정과 열 너비까지 맞춰 바로 볼 수 있게 한다."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("Excel 내보내기에는 openpyxl 이 필요합니다: pip install openpyxl") from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "리뷰 분석"

    headers = [header for _, header in COLUMNS]
    worksheet.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E86DE")
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for record in records:
        worksheet.append([record.get(header) for header in headers])

    widths = {"ID": 6, "작성일": 12, "제품": 20, "카테고리": 10, "별점": 6, "감정": 10,
              "신뢰도": 8, "키워드": 24, "언어": 6, "글자수": 8, "리뷰내용": 60,
              "분석모델": 14, "분석일시": 20}
    for index, header in enumerate(headers, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = widths.get(header, 14)

    worksheet.freeze_panes = "A2"  # 헤더 고정
    workbook.save(path)
    return path


def export_reviews(
    db: Database, cfg: dict[str, Any], fmt: str = "csv",
    output: str | Path | None = None, **filters,
) -> tuple[Path, int]:
    """조건에 맞는 리뷰를 내보낸다. 반환: (저장 경로, 건수)"""
    rows = db.query_reviews(sort="date", order="desc", **filters)
    records = rows_to_dicts(rows)

    if output:
        path = Path(output)
    else:
        exports_dir = config_module.resolve_path(cfg["paths"]["exports"])
        exports_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = exports_dir / f"reviews_{stamp}.{fmt}"

    path.parent.mkdir(parents=True, exist_ok=True)

    writers = {"csv": export_csv, "jsonl": export_jsonl, "xlsx": export_xlsx}
    if fmt not in writers:
        raise ValueError(f"지원하지 않는 형식입니다: {fmt} (csv, jsonl, xlsx)")

    writers[fmt](records, path)
    return path, len(records)


# --------------------------------------------------------------------- CLI

def cmd_export(args, cfg: dict[str, Any]) -> int:
    filters: dict[str, Any] = {}
    for attr in ("sentiment", "rating_min", "rating_max", "date_from", "date_to", "product"):
        value = getattr(args, attr, None)
        if value is not None:
            filters[attr] = value

    db_path = config_module.resolve_path(cfg["paths"]["db"])
    with Database(db_path) as db:
        if db.count_reviews(**filters) == 0:
            logger.warning("조건에 맞는 리뷰가 없어 내보낼 데이터가 없습니다.")
            return 1
        path, count = export_reviews(db, cfg, fmt=args.format, output=args.output, **filters)

    described = ", ".join(f"{k}={v}" for k, v in filters.items()) or "전체"
    logger.info("내보내기 완료: %d건 (%s) → %s", count, described, path)
    return 0
