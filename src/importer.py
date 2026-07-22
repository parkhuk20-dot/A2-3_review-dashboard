"""리뷰 데이터 수집: CSV/Excel 파일 읽기(import)와 수동 1건 추가(add).

원본은 손대지 않고 raw_reviews 에 그대로 넣는다. 검증·정규화는 clean 단계 몫이다.
파일마다 헤더 이름이 제각각이라(리뷰내용/review_text/content …) 별칭 표로 매핑한다.
"""

from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import Any, Iterator

from src import config as config_module
from src.db import Database

logger = logging.getLogger(__name__)

CSV_SUFFIXES = {".csv", ".tsv", ".txt"}
EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx"}


def _normalize_header(name: str) -> str:
    """헤더 비교용 정규화: 소문자 + 공백/밑줄/하이픈 제거."""
    return str(name or "").strip().lower().replace(" ", "").replace("_", "").replace("-", "")


def build_column_map(headers: list[str], aliases: dict[str, list[str]]) -> dict[str, str]:
    """파일 헤더를 내부 필드명으로 매핑한다.

    반환: {내부필드: 실제헤더}. 매칭 안 된 필드는 빠진다.
    """
    normalized = {_normalize_header(h): h for h in headers if h is not None}
    column_map: dict[str, str] = {}

    for field, alias_list in aliases.items():
        for alias in alias_list:
            key = _normalize_header(alias)
            if key in normalized:
                column_map[field] = normalized[key]
                break

    return column_map


def read_csv_rows(path: Path, encoding: str) -> Iterator[dict[str, Any]]:
    """CSV/TSV 를 행 딕셔너리로 읽는다."""
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with path.open("r", encoding=encoding, newline="") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        if not reader.fieldnames:
            return
        for row in reader:
            yield {k: v for k, v in row.items() if k is not None}


def read_excel_rows(path: Path, sheet: str | None = None) -> Iterator[dict[str, Any]]:
    """Excel 시트를 행 딕셔너리로 읽는다. 첫 행을 헤더로 본다."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - 설치 안내용
        raise RuntimeError("Excel 을 읽으려면 openpyxl 이 필요합니다: pip install openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet] if sheet else workbook[workbook.sheetnames[0]]

    rows = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        workbook.close()
        return

    headers = [str(h) if h is not None else "" for h in header_row]
    for values in rows:
        yield {headers[i]: values[i] for i in range(min(len(headers), len(values)))}

    workbook.close()


def _stringify(value: Any) -> str | None:
    """엑셀 날짜(datetime)·숫자를 문자열로 통일한다. 빈 값은 None."""
    if value is None:
        return None
    if hasattr(value, "strftime"):  # datetime / date
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def import_file(
    db: Database,
    file_path: str | Path,
    cfg: dict[str, Any],
    sheet: str | None = None,
    encoding: str | None = None,
    limit: int | None = None,
    product_override: str | None = None,
    category_override: str | None = None,
) -> dict[str, int]:
    """파일을 읽어 raw_reviews 에 저장하고 집계를 반환한다."""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {path}")

    suffix = path.suffix.lower()
    import_cfg = cfg.get("import", {})
    aliases = import_cfg.get("column_aliases", {})

    if suffix in EXCEL_SUFFIXES:
        rows = read_excel_rows(path, sheet)
    elif suffix in CSV_SUFFIXES:
        rows = read_csv_rows(path, encoding or import_cfg.get("default_encoding", "utf-8-sig"))
    else:
        raise ValueError(f"지원하지 않는 파일 형식입니다: {suffix} (csv, tsv, xlsx 지원)")

    logger.info("파일 로드: %s", path)

    total = inserted = skipped = 0
    column_map: dict[str, str] | None = None

    for row_no, row in enumerate(rows, start=2):  # 2 = 헤더 다음 행
        if limit is not None and inserted >= limit:
            break

        if column_map is None:
            column_map = build_column_map(list(row.keys()), aliases)
            if "review_text" not in column_map:
                raise ValueError(
                    "리뷰 텍스트 컬럼을 찾지 못했습니다. "
                    f"파일 헤더: {list(row.keys())} / "
                    f"인식 가능한 이름: {aliases.get('review_text', [])}"
                )
            logger.info("컬럼 매핑: %s", column_map)

        total += 1
        review_text = _stringify(row.get(column_map["review_text"]))

        # 리뷰 본문이 비어 있으면 저장할 가치가 없어 여기서 거른다.
        if not review_text:
            skipped += 1
            logger.warning("%d행: 리뷰 텍스트가 비어 있어 건너뜁니다.", row_no)
            continue

        record = {
            "source_file": str(path),
            "source_row": row_no,
            "review_text": review_text,
            "rating": _stringify(row.get(column_map.get("rating", ""), None)),
            "review_date": _stringify(row.get(column_map.get("review_date", ""), None)),
            "product": _stringify(row.get(column_map.get("product", ""), None)) or product_override,
            "category": (_stringify(row.get(column_map.get("category", ""), None))
                         or category_override),
            "raw": {k: _stringify(v) for k, v in row.items()},
        }
        db.insert_raw(record)
        inserted += 1

    if column_map is None:
        logger.warning("읽을 데이터가 없습니다: %s", path)

    db.log_import(str(path), total, inserted, skipped)
    return {"total": total, "inserted": inserted, "skipped": skipped}


# --------------------------------------------------------------------- CLI

def cmd_import(args, cfg: dict[str, Any]) -> int:
    db_path = config_module.resolve_path(cfg["paths"]["db"])
    with Database(db_path) as db:
        result = import_file(
            db, args.file, cfg,
            sheet=args.sheet, encoding=args.encoding, limit=args.limit,
            product_override=args.product, category_override=args.category,
        )

    logger.info(
        "총 %d건 감지, 유효 %d건, 스킵 %d건 (본문 없음)",
        result["total"], result["inserted"], result["skipped"],
    )
    logger.info("raw 저장소에 저장 완료 → %s", db_path)
    logger.info("다음 단계: python main.py clean")
    return 0


def cmd_add(args, cfg: dict[str, Any]) -> int:
    """파일 없이 리뷰 1건을 직접 추가한다."""
    text = (args.text or "").strip()
    if not text:
        logger.error("리뷰 본문(--text)이 비어 있습니다.")
        return 2

    db_path = config_module.resolve_path(cfg["paths"]["db"])
    with Database(db_path) as db:
        raw_id = db.insert_raw({
            "source_file": "(수동 입력)",
            "source_row": None,
            "review_text": text,
            "rating": args.rating,
            "review_date": args.review_date,
            "product": args.product,
            "category": args.category,
            "raw": {
                "review_text": text, "rating": args.rating,
                "review_date": args.review_date, "product": args.product,
                "category": args.category,
            },
        })
        db.log_import("(수동 입력)", 1, 1, 0)

    logger.info("리뷰를 raw 저장소에 추가했습니다 (raw_id=%d)", raw_id)
    logger.info("다음 단계: python main.py clean")
    return 0
